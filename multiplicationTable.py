# Takes a number N from the comand line and creates an N by N multiplication table in excel.

import openpyxl, sys, pyinputplus
from openpyxl.utils import get_column_letter

# Create a workbook.
print('Creating a workbook...')
wb = openpyxl.Workbook()
sheet = wb['Sheet']  # We will be working in the default sheet 'Sheet'.
                    
# Add function to get arguements from command line.
if len(sys.argv) == 2:
    user_number = sys.argv[1]
else:
    user_number = pyinputplus.inputInt('Enter an integer for max row/column of multiplication table: ')

# Enter the number 1 through N into cells 1st column and 1st row
for i in range(1, user_number + 1):
    sheet['A' + str(i+1)].value = i  # Starting at the second row
    sheet[get_column_letter(i+1) + '1'].value = i
    

# Loop through filling out the table
for i in range(2, user_number + 2):
    curr_col = get_column_letter(i)  # Get the current column number
    for j in range(2, user_number + 2):  # Fill in the row
        sheet[curr_col + str(j)].value = sheet[curr_col + '1'].value * sheet['A' + str(j)].value


# Save the file as multiplicationTable.xlsx
sheet.freeze_panes = 'A2'  # Freezes the top row so  you can see it even when scrolling down
wb.save('multiplicationTable.xlsx')
print('Done')
